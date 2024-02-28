from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Optional, NamedTuple, Tuple

import openpyxl as pyxl
import openpyxl.writer.excel
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

from merger._config import ConfigProperty, Config
from merger.exceptions import MergeException


class Merger:

    def __init__(self, original_file_path: Path, new_file_path: Path, column_key: str,
                 merged_file_name: Optional[str] = None):
        """
        Given two spreadsheet files, and a column name that is used to identify each row, merge the two spreadsheet
        files together. If there are any rows that are shared between both spreadsheets, the new the original
        spreadsheet's cells will be updated based on the information contained in the new spreadsheet. If there are
        rows in the new spreadsheet that do not exist in the original, they will also be added to the original sheet.
        Then, these merged sheet is either saved/overwritten to the original file, or is saved to a different file.
        """
        self.original_file_path = original_file_path.resolve()
        self.new_file_path = new_file_path.resolve()
        self.column_key = column_key

        if self.new_file_path == self.original_file_path:
            raise MergeException("Spreadsheet file paths cannot be identical.")

        # If the merged file name is blank, assume that the main file will be used instead
        if not merged_file_name:
            self.merged_file_path = self.original_file_path
        else:
            merged_file_dir = self.original_file_path.parent
            merged_file_ext = self.original_file_path.suffix
            self.merged_file_path = merged_file_dir.joinpath(f"{merged_file_name}{merged_file_ext}").resolve()

        self._original_wb: Workbook = pyxl.load_workbook(self.original_file_path)
        self._original_wb.active = 0
        self._original_sheet: Worksheet = self._original_wb.active
        self._original_max_row: int = self._original_sheet.max_row

        self._new_wb: Workbook = pyxl.load_workbook(self.new_file_path)
        self._new_wb.active = 0
        self._new_sheet: Worksheet = self._new_wb.active
        self._new_max_row: int = self._new_sheet.max_row

        # Close workbooks, the files no longer need to stay open
        self._clean_stop()

        # Locate header rows
        self.original_header_row = self._locate_header_row(self._original_sheet)
        self._original_first_data_row = self.original_header_row + 1
        self.new_header_row = self._locate_header_row(self._new_sheet)
        self._new_first_data_row = self.new_header_row + 1
        if not (self.original_header_row and self.new_header_row):
            raise MergeException(f"The key '{self.column_key}' is not a column label in either of the spreadsheets.")
        elif not self.original_header_row:
            raise MergeException(f"The key '{self.column_key}' is not a column label in `{self.original_file_path}`.")
        elif not self.new_header_row:
            raise MergeException(f"The key '{self.column_key}' is not a column label in `{self.new_file_path}`.")

        # Locate column label positions
        self._label_indices = self._locate_labels()

        # Locate first row of data to use for formatting later
        self._format_row: Tuple[Cell] = next(self._original_sheet.iter_rows(min_row=self._original_first_data_row,
                                                                            max_row=self._original_first_data_row))

    def merge(self):
        try:
            self._hook_initialization()
            key_col_indices = self._label_indices[self.column_key]

            key_rows = self._map_main_keys(key_col_indices)

            # Find all rows in the main sheet based on the current key value in the new sheet
            new_rows = self._new_sheet.iter_rows(min_row=self._new_first_data_row, max_row=self._new_max_row)
            for new_key_row_index, new_key_row in enumerate(new_rows):
                new_key_val = new_key_row[key_col_indices.new_label_index].value

                # If a row wasn't found for this key in the main sheet, it must be new and inserted into the sheet
                if new_key_val not in key_rows:
                    self._original_sheet.insert_rows(self._original_first_data_row)
                    main_key_row = next(self._original_sheet.iter_rows(min_row=self._original_first_data_row,
                                                                       max_row=self._original_first_data_row))

                    self._update_row_formatting(main_key_row)
                else:
                    main_key_row = key_rows[new_key_val]

                # Update all previous keys in the main sheet with new sheet data
                self._update_row(main_key_row, new_key_row)

                self._hook_row_merged(new_key_row_index)

            self._hook_pre_saving()

            self._add_timestamp()

            # Creates a copy
            openpyxl.writer.excel.save_workbook(self._original_wb, str(self.merged_file_path))
            self._hook_success()
        except BaseException as e:
            self._hook_exception()
            raise e from e
        finally:
            self._clean_stop()

    def _clean_stop(self):
        # Close workbook files
        self._original_wb.close()
        self._new_wb.close()

    class _LabelIndices(NamedTuple):
        main_label_index: int
        new_label_index: Optional[int]

    def _locate_header_row(self, sheet) -> int:
        # Iterate over all columns in the main sheet to find the header column (based on where the column key is)
        for row in sheet.rows:
            for cell in row:
                if cell.value == self.column_key:
                    return cell.row

        return 0

    def _locate_labels(self):
        """Map column label indexes used in new sheet to the main sheet indexes"""
        label_indices: dict[str, Merger._LabelIndices] = {}

        # Iterate over all the column labels in the main sheet
        for (main_label_cell,) in self._original_sheet.iter_cols(min_row=self.original_header_row,
                                                                 max_row=self.original_header_row):
            # if the main column label is empty, skip it
            if main_label_cell.value is None:
                continue

            # Find new sheet position of current column label being examined in the main sheet
            for (new_label_cell,) in self._new_sheet.iter_cols(min_row=self.new_header_row,
                                                               max_row=self.new_header_row):

                # if the new column label is empty, skip it
                if new_label_cell.value is None:
                    continue

                # Keep track of where the label exists in each
                if main_label_cell.value.lower() == new_label_cell.value.lower():
                    label_indices[main_label_cell.value] = self._LabelIndices(main_label_cell.column - 1,
                                                                              new_label_cell.column - 1)
                    break

            # Column label in main does not exist in new, and should be ignored later on
            if main_label_cell.value not in label_indices:
                label_indices[main_label_cell.value] = self._LabelIndices(main_label_cell.column, None)

        return label_indices

    def _map_main_keys(self, key_col_indices):
        main_key_rows = {}
        for key_row_index, key_row in enumerate(self._original_sheet.iter_rows(min_row=self._original_first_data_row)):
            main_key_val = key_row[key_col_indices.main_label_index].value
            main_key_rows[main_key_val] = key_row

            self._hook_row_indexed(key_row_index)

        return main_key_rows

    def _update_row(self, main_row: Tuple[Cell], new_row: Tuple[Cell]):
        for indexes in self._label_indices.values():
            # If label does not exist in the new sheet, just ignore this label
            if indexes.new_label_index is None:
                continue

            main_row[indexes.main_label_index].value = new_row[indexes.new_label_index].value

    def _update_row_formatting(self, row: Tuple[Cell]):
        for cell_index, cell in enumerate(row):
            format_cell: Cell = self._format_row[cell_index]

            cell.font = copy(format_cell.font)
            cell.fill = copy(format_cell.fill)
            cell.border = copy(format_cell.border)
            cell.number_format = copy(format_cell.number_format)
            cell.protection = copy(format_cell.protection)
            cell.alignment = copy(format_cell.alignment)
            cell.quotePrefix = copy(format_cell.quotePrefix)
            cell.pivotButton = copy(format_cell.pivotButton)

    def _add_timestamp(self):
        timestamp_cell_str = Config.get(ConfigProperty.MERGE_TIMESTAMP_CELL)

        # Add a row in case the header row and timestamp cell might overlap
        if self._original_sheet[timestamp_cell_str].row == self.original_header_row:
            self._original_sheet.insert_rows(1)

        # Update cell to indicate the date the sheet was modified
        modified_time_str = datetime.now().strftime("%A %B %d %Y, %I:%M%p")
        self._original_sheet[Config.get(ConfigProperty.MERGE_TIMESTAMP_CELL)].value = modified_time_str


    def _hook_initialization(self):
        """
        Hook that runs immediately before both the indexing and merge processes are about to start.
        """

    def _hook_row_indexed(self, row_index):
        """
        Hook that runs when a row has been indexed, but not yet merged.

        Args:
            row_index (int): Index of the row that was just indexed pre-merge.
        """

    def _hook_row_merged(self, row_index):
        """
        Hook that runs immediately after a row in the spreadsheet has been merged.

        Args:
            row_index (int): Index of the row that was just merged.
        """

    def _hook_pre_saving(self):
        """
        Hook that runs right before saving the merged rows to a new file.
        """

    def _hook_success(self):
        """
        Hook that runs when the merge process completes successfully.
        """

    def _hook_exception(self):
        """
        Hook that runs when the merge process fails with an exception.
        """
